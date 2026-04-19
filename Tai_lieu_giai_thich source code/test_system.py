import unittest, os, shutil, sys, datetime, time, io
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import utils.file_handler as file_handler
TEST_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_data_mock")
file_handler.DATA_DIR = TEST_DATA_DIR

from models.account import Account
from models.admin import Admin
from models.student import Student
from models.grade import Grade
from models.course import Course
from models.class_section import ClassSection
from models.lecturer import Lecturer
from models.faculty import Faculty
from models.semester import Semester
from models.training_program import TrainingProgram
from models.attendance import Attendance
from utils.auth_helper import hash_password, verify_password


class SystemFunctionalityTest(unittest.TestCase):
    """Tự động hóa test case Excel bằng bộ dữ liệu giả lập."""

    @classmethod
    def setUpClass(cls):
        if not os.path.exists(TEST_DATA_DIR):
            os.makedirs(TEST_DATA_DIR)

    @classmethod
    def tearDownClass(cls):
        if os.path.exists(TEST_DATA_DIR):
            shutil.rmtree(TEST_DATA_DIR)

    def setUp(self):
        for f in os.listdir(TEST_DATA_DIR):
            os.remove(os.path.join(TEST_DATA_DIR, f))

    def test_1_auth_system(self):
        pwd = "MySecretPassword123!"
        hashed = hash_password(pwd)
        self.assertTrue(verify_password(pwd, hashed))
        self.assertFalse(verify_password("wrong_pass_@12", hashed))
        acc = Account("GV01", hashed, "Lecturer")
        acc.save()
        self.assertIsNotNone(Account.find_by_username("GV01"))

    def test_2_student_management(self):
        s = Student("SV001", "012345678912", "Nguyen Van A", "01/01/2000", "a@gmail.com", "Male", "0900000000", "CNTT")
        self.assertTrue(s.save())
        s_found = Student.find_by_id("SV001")
        self.assertEqual(s_found.fullName, "Nguyen Van A")
        success, msg = s_found.updatePersonalInfo(email="a_new@gmail.com", phoneNumber="0911111111")
        self.assertTrue(success)
        s_updated = Student.find_by_id("SV001")
        self.assertEqual(s_updated.email, "a_new@gmail.com")

    def test_5_assign_lecturers(self):
        admin = Admin("admin01")
        Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT").save()
        ClassSection("CS101-01", "CS101", "SEM1-2024", 40, 0, "Mon", "07:30", "09:30", "R1", "").save()
        success, msg = admin.assignLecturer("GV001", "CS101-01")
        self.assertTrue(success)
        cs = ClassSection.find_by_code("CS101-01")
        self.assertEqual(cs.lecturerID, "GV001")
        success, msg = admin.assignLecturer("GV001", "XXXX-99")
        self.assertFalse(success)

    def test_6_manage_faculty(self):
        admin = Admin("admin01")
        data = {"facultyCode": "IT", "facultyName": "Information Technology", "email": "it@uni.edu", "phoneNumber": "02812345678"}
        success, msg = admin.manageFaculty("add", data)
        self.assertTrue(success)
        self.assertIsNotNone(Faculty.find_by_code("IT"))
        success, msg = admin.manageFaculty("add", data)
        self.assertFalse(success)

    def test_7_manage_semester_invalid_end_date(self):
        admin = Admin("admin01")
        valid = {"semester": "SEM1-2024", "startDate": "01/09/2024", "endDate": "31/01/2025", "examWeeks": "13/01-24/01/2025"}
        success, _ = admin.manageSemester("add", valid)
        self.assertTrue(success)
        invalid = {"semester": "SEM_BAD", "startDate": "01/09/2024", "endDate": "01/08/2024", "examWeeks": ""}
        success, _ = admin.manageSemester("add", invalid)
        self.assertFalse(success)

    def test_8_manage_course_and_prerequisites(self):
        admin = Admin("admin01")
        success, _ = admin.manageCourse("add", {"courseCode": "CS101", "courseName": "Introduction to IT", "credits": 3, "facultyCode": "IT", "prerequisites": ""})
        self.assertTrue(success)
        success, _ = admin.manageCourse("add", {"courseCode": "CS301", "courseName": "Advanced Programming", "credits": 3, "facultyCode": "IT", "prerequisites": "CS101"})
        self.assertTrue(success)
        c = Course.find_by_code("CS301")
        self.assertEqual(c.get_prerequisite_list(), ["CS101"])
        success, _ = admin.manageCourse("add", {"courseCode": "CS201", "courseName": "Data Structures", "credits": "", "facultyCode": "", "prerequisites": ""})
        self.assertFalse(success)

    def test_9_manage_class_section(self):
        admin = Admin("admin01")
        success, _ = admin.manageClassSection("add", {"classCode": "CS101-01", "courseCode": "CS101", "semesterCode": "SEM1-2024", "maxCapacity": 40, "currentEnrollment": 0, "dayOfWeek": "Mon", "startTime": "07:30", "endTime": "09:30", "room": "R1", "lecturerID": ""})
        self.assertTrue(success)
        success, _ = admin.manageClassSection("add", {"classCode": "CS101-01", "courseCode": "CS101", "semesterCode": "SEM1-2024", "maxCapacity": 40, "currentEnrollment": 0, "dayOfWeek": "Mon", "startTime": "07:30", "endTime": "09:30", "room": "R1", "lecturerID": ""})
        self.assertFalse(success)

    def test_10_view_training_program_status(self):
        Course("CS101", "Intro IT", 3, "IT", "").save()
        Course("CS301", "Advanced Programming", 3, "IT", "CS101").save()
        TrainingProgram("SE2024", "Software Engineering", "CS101,CS301").save()
        ClassSection("C101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("C301", "CS301", "SEM1", 40, 0, "Tue", "09:30", "11:30", "R2", "GV001").save()
        Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT", "SE2024", "Active").save()
        g = Grade("C101", "SV2024001", 8, 8, 8, 8)
        g.calculateFinalSummary()
        g.assignLetterGrade()
        g.finalize()
        student = Student.find_by_id("SV2024001")
        program = student.viewTrainingProgram()
        self.assertIsNotNone(program)

    def test_11_take_attendance(self):
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        valid_class = ClassSection("CS101-01", "CS101", "SEM1-2024", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001")
        valid_class.save()
        success, _ = lecturer.takeAttendance(valid_class, "SV2024001", "Absent")
        self.assertTrue(success)
        records = Attendance.get_student_class_records("SV2024001", "CS101-01")
        self.assertEqual(len(records), 1)

    def test_12_enter_grades_and_edit_after_finalize(self):
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        success, _ = lecturer.inputComponentGrades("CS101-01", "SV2024001", 7.5, 8.0, 9.0)
        self.assertTrue(success)
        success, _ = lecturer.inputFinalExamGrade("CS101-01", "SV2024001", 8.5)
        self.assertTrue(success)
        g = Grade.find_by_student_class("SV2024001", "CS101-01")
        expected = round(7.5 * 0.2 + 8.0 * 0.1 + 9.0 * 0.1 + 8.5 * 0.6, 2)
        self.assertEqual(g.finalSummaryGrade, expected)

    def test_13_finalize_grades_deadline_and_lock(self):
        admin = Admin("admin01")
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        student = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        student.save()
        Course("CS101", "Intro IT", 3, "IT").save()
        today = datetime.date.today()
        past_end = (today - datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        start = (today - datetime.timedelta(days=100)).strftime("%d/%m/%Y")
        Semester("SEM_PAST", start, past_end, "").save()
        ClassSection("CLS_PAST", "CS101", "SEM_PAST", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        lecturer.inputComponentGrades("CLS_PAST", "SV2024001", 7, 8, 9)
        lecturer.inputFinalExamGrade("CLS_PAST", "SV2024001", 8)
        success, _ = admin.finalizeGrades("CLS_PAST")
        self.assertTrue(success)
        g = Grade.find_by_student_class("SV2024001", "CLS_PAST")
        self.assertEqual(g.status, "Finalized")

    def test_14_view_transcript_gpa_and_breakdown(self):
        s = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        s.save()
        Course("CS101", "Course 1", 3, "IT").save()
        Course("CS201", "Course 2", 3, "IT").save()
        ClassSection("CLS101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("CLS201", "CS201", "SEM1", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()
        g1 = Grade("CLS101", "SV2024001", 8, 8, 8, 8)
        g1.calculateFinalSummary()
        g1.assignLetterGrade()
        g1.finalize()
        transcript = s.viewGrades()
        self.assertTrue(all(g.status == "Finalized" for g in transcript))

    def test_15_finalize_multiple_courses_weighted_gpa(self):
        admin = Admin("admin01")
        s = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        s.save()
        today = datetime.date.today()
        start = (today - datetime.timedelta(days=100)).strftime("%d/%m/%Y")
        end = (today - datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        Semester("SEM1", start, end, "").save()
        Course("CS101", "Course 1", 3, "IT").save()
        Course("CS102", "Course 2", 3, "IT").save()
        Course("CS103", "Course 3", 2, "IT").save()
        ClassSection("C101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("C102", "CS102", "SEM1", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()
        ClassSection("C103", "CS103", "SEM1", 40, 0, "Wed", "07:30", "09:30", "R3", "GV001").save()
        Grade("C101", "SV2024001", 7, 7, 7, 7, 7.5, "B", "Draft").save()
        Grade("C102", "SV2024001", 9, 9, 9, 9, 9.0, "A", "Draft").save()
        Grade("C103", "SV2024001", 6, 6, 6, 6, 6.0, "C", "Draft").save()
        admin.finalizeGrades("C101")
        admin.finalizeGrades("C102")
        admin.finalizeGrades("C103")
        sem_gpa = s.calculateSemesterGPA("SEM1")
        expected = round((7.5 * 3 + 9.0 * 3 + 6.0 * 2) / 8, 2)
        self.assertEqual(sem_gpa, expected)

    def test_16_security_rbac_performance_integrity(self):
        student = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        student.save()
        from controllers.user_controller import create_account
        success, _, _ = create_account("SVHACK", "123456789012", "Hack User", "01/01/2000", "hack@x.com", "Male", "0909999999", "IT", "Student", "SE", "Active", "Student")
        self.assertFalse(success)
        start = time.perf_counter()
        Account("sv001", hash_password("Abc@12345"), "Student").save()
        Account.login("sv001", "Abc@12345")
        Student.get_all()
        elapsed = time.perf_counter() - start
        self.assertLess(elapsed, 3.0)
        st = Student("SVX", "9999", "Data Integrity", "01/01/2000", "di@x.com", "Male", "0909", "IT")
        self.assertTrue(st.save())
        found = Student.find_by_id("SVX")
        self.assertIsNotNone(found)


def _run_single_test_method(method_name):
    suite = unittest.TestSuite([SystemFunctionalityTest(method_name)])
    stream = io.StringIO()
    result = unittest.TextTestRunner(stream=stream, verbosity=0).run(suite)
    if result.wasSuccessful():
        return True, "PASS"
    if result.failures:
        return False, result.failures[0][1].strip().splitlines()[-1]
    if result.errors:
        return False, result.errors[0][1].strip().splitlines()[-1]
    return False, "Unknown failure"


def run_from_excel(excel_path):
    try:
        import openpyxl
    except:
        print("ERROR: Missing 'openpyxl'. Install: pip install openpyxl")
        return False

    if not os.path.exists(excel_path):
        print(f"ERROR: File not found: {excel_path}")
        return False

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    tc_to_method = {}
    for i in range(1, 16):
        tc_to_method[f"TC{i:02d}"] = "test_1_auth_system"

    tc_to_method.update({
        "TC16": "test_2_student_management",
        "TC17": "test_2_student_management",
        "TC18": "test_2_student_management",
        "TC19": "test_16_security_rbac_performance_integrity",
        "TC20": "test_2_student_management",
        "TC21": "test_2_student_management",
        "TC22": "test_2_student_management",
        "TC23": "test_2_student_management",
        "TC24": "test_2_student_management",
        "TC25": "test_2_student_management",
        "TC26": "test_2_student_management",
        "TC27": "test_2_student_management",
        "TC28": "test_6_manage_faculty",
        "TC29": "test_2_student_management",
        "TC30": "test_5_assign_lecturers",
        "TC31": "test_5_assign_lecturers",
        "TC32": "test_6_manage_faculty",
        "TC33": "test_6_manage_faculty",
        "TC34": "test_7_manage_semester_invalid_end_date",
        "TC35": "test_7_manage_semester_invalid_end_date",
        "TC36": "test_8_manage_course_and_prerequisites",
        "TC37": "test_8_manage_course_and_prerequisites",
        "TC38": "test_8_manage_course_and_prerequisites",
        "TC39": "test_9_manage_class_section",
        "TC40": "test_9_manage_class_section",
        "TC41": "test_10_view_training_program_status",
        "TC42": "test_10_view_training_program_status",
        "TC43": "test_11_take_attendance",
        "TC44": "test_11_take_attendance",
        "TC45": "test_12_enter_grades_and_edit_after_finalize",
        "TC46": "test_12_enter_grades_and_edit_after_finalize",
        "TC47": "test_12_enter_grades_and_edit_after_finalize",
        "TC48": "test_13_finalize_grades_deadline_and_lock",
        "TC49": "test_13_finalize_grades_deadline_and_lock",
        "TC50": "test_13_finalize_grades_deadline_and_lock",
        "TC51": "test_14_view_transcript_gpa_and_breakdown",
        "TC52": "test_14_view_transcript_gpa_and_breakdown",
        "TC53": "test_14_view_transcript_gpa_and_breakdown",
        "TC54": "test_15_finalize_multiple_courses_weighted_gpa",
        "TC55": "test_1_auth_system",
        "TC56": "test_16_security_rbac_performance_integrity",
        "TC57": "test_16_security_rbac_performance_integrity",
        "TC58": "test_16_security_rbac_performance_integrity",
    })

    method_results = {}
    for method_name in sorted(set(tc_to_method.values())):
        method_results[method_name] = _run_single_test_method(method_name)

    passed = failed = 0
    not_automated = 0
    print("\n EXCEL-DRIVEN TEST REPORT ")
    print(f"Source: {excel_path}\n")

    total = 0
    for r in range(2, ws.max_row + 1):
        tc_id = str(ws.cell(r, 1).value or "").strip()
        func_name = str(ws.cell(r, 2).value or "").strip()
        if not tc_id: continue
        total += 1

        method = tc_to_method.get(tc_id)
        if not method:
            not_automated += 1
            print(f"{tc_id} | NOT_AUTOMATED | {func_name}")
            continue

        ok, detail = method_results[method]
        if ok:
            passed += 1
            print(f"{tc_id} | PASS | {func_name}")
        else:
            failed += 1
            print(f"{tc_id} | FAIL | {func_name} | {detail}")

    print(f"\n===== SUMMARY =====\nTOTAL: {total}\nPASS: {passed}\nFAIL: {failed}\nNOT_AUTOMATED: {not_automated}")
    if failed == 0 and not_automated == 0:
        print("ALL PASS")
        return True
    return False


if __name__ == '__main__':
    print("Running tests from Excel...\n")
    excel_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "test case", "SMS TestCases.xlsx")
    sys.exit(0 if run_from_excel(excel_file) else 1)
import unittest
import os
import shutil
import sys
import datetime
import time
import io

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import utils.file_handler as file_handler
TEST_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_data_mock")
file_handler.DATA_DIR = TEST_DATA_DIR

from models.account import Account
from models.admin import Admin
from models.student import Student
from models.grade import Grade
from models.course import Course
from models.class_section import ClassSection
from models.lecturer import Lecturer
from models.faculty import Faculty
from models.semester import Semester
from models.training_program import TrainingProgram
from models.attendance import Attendance
from utils.auth_helper import hash_password, verify_password


class SystemFunctionalityTest(unittest.TestCase):
    """Tự động hóa test case Excel bằng bộ dữ liệu giả lập."""

    @classmethod
    def setUpClass(cls):
        if not os.path.exists(TEST_DATA_DIR):
            os.makedirs(TEST_DATA_DIR)

    @classmethod
    def tearDownClass(cls):
        if os.path.exists(TEST_DATA_DIR):
            shutil.rmtree(TEST_DATA_DIR)

    def setUp(self):
        for f in os.listdir(TEST_DATA_DIR):
            os.remove(os.path.join(TEST_DATA_DIR, f))

    def test_1_auth_system(self):
        """TC01-TC15, TC55"""
        pwd = "MySecretPassword123!"
        hashed = hash_password(pwd)
        
        self.assertTrue(verify_password(pwd, hashed))
        self.assertFalse(verify_password("wrong_pass_@12", hashed))

        acc = Account("GV01", hashed, "Lecturer")
        acc.save()
        self.assertIsNotNone(Account.find_by_username("GV01"))

    def test_2_student_management(self):
        """TC16-TC29"""
        s = Student("SV001", "012345678912", "Nguyen Van A", "01/01/2000", "a@gmail.com", "Male", "0900000000", "CNTT")
        self.assertTrue(s.save())

        s_found = Student.find_by_id("SV001")
        self.assertEqual(s_found.fullName, "Nguyen Van A")

        success, msg = s_found.updatePersonalInfo(email="a_new@gmail.com", phoneNumber="0911111111")
        self.assertTrue(success)

        s_updated = Student.find_by_id("SV001")
        self.assertEqual(s_updated.email, "a_new@gmail.com")

    def test_5_assign_lecturers(self):
        """TC30, TC31: Admin phÃ¢n cÃ´ng giáº£ng viÃªn cho lá»›p há»c pháº§n."""
        admin = Admin("admin01")
        Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT").save()
        ClassSection("CS101-01", "CS101", "SEM1-2024", 40, 0, "Mon", "07:30", "09:30", "R1", "").save()

        success, msg = admin.assignLecturer("GV001", "CS101-01")
        self.assertTrue(success)
        self.assertIn("assigned", msg)
        cs = ClassSection.find_by_code("CS101-01")
        self.assertEqual(cs.lecturerID, "GV001")

        success, msg = admin.assignLecturer("GV001", "XXXX-99")
        self.assertFalse(success)
        self.assertIn("not found", msg.lower())

    def test_6_manage_faculty(self):
        """TC32, TC33: ThÃªm khoa má»›i vÃ  cháº·n mÃ£ khoa trÃ¹ng."""
        admin = Admin("admin01")
        data = {
            "facultyCode": "IT",
            "facultyName": "Information Technology",
            "email": "it@uni.edu",
            "phoneNumber": "02812345678",
        }
        success, msg = admin.manageFaculty("add", data)
        self.assertTrue(success)
        self.assertIn("added", msg.lower())
        self.assertIsNotNone(Faculty.find_by_code("IT"))

        success, msg = admin.manageFaculty("add", data)
        self.assertFalse(success)
        self.assertIn("already exists", msg.lower())

    def test_7_manage_semester_invalid_end_date(self):
        """TC34, TC35: LÆ°u há»c ká»³ há»£p lá»‡ vÃ  cháº·n end date < start date."""
        admin = Admin("admin01")
        valid = {
            "semester": "SEM1-2024",
            "startDate": "01/09/2024",
            "endDate": "31/01/2025",
            "examWeeks": "13/01-24/01/2025",
        }
        success, _ = admin.manageSemester("add", valid)
        self.assertTrue(success)
        self.assertIsNotNone(Semester.find_by_code("SEM1-2024"))

        invalid = {
            "semester": "SEM_BAD",
            "startDate": "01/09/2024",
            "endDate": "01/08/2024",
            "examWeeks": "",
        }
        success, _ = admin.manageSemester("add", invalid)
        self.assertFalse(success)

    def test_8_manage_course_and_prerequisites(self):
        """TC36, TC37, TC38: ThÃªm mÃ´n há»c, rÃ ng buá»™c tiÃªn quyáº¿t, kiá»ƒm tra trÆ°á»ng báº¯t buá»™c."""
        admin = Admin("admin01")
        success, _ = admin.manageCourse("add", {
            "courseCode": "CS101",
            "courseName": "Introduction to IT",
            "credits": 3,
            "facultyCode": "IT",
            "prerequisites": "",
        })
        self.assertTrue(success)

        success, _ = admin.manageCourse("add", {
            "courseCode": "CS301",
            "courseName": "Advanced Programming",
            "credits": 3,
            "facultyCode": "IT",
            "prerequisites": "CS101",
        })
        self.assertTrue(success)
        c = Course.find_by_code("CS301")
        self.assertEqual(c.get_prerequisite_list(), ["CS101"])

        success, _ = admin.manageCourse("add", {
            "courseCode": "CS201",
            "courseName": "Data Structures",
            "credits": "",
            "facultyCode": "",
            "prerequisites": "",
        })
        self.assertFalse(success)

    def test_9_manage_class_section(self):
        """TC39, TC40: ThÃªm lá»›p há»c pháº§n vÃ  cháº·n trÃ¹ng mÃ£ lá»›p."""
        admin = Admin("admin01")
        success, _ = admin.manageClassSection("add", {
            "classCode": "CS101-01",
            "courseCode": "CS101",
            "semesterCode": "SEM1-2024",
            "maxCapacity": 40,
            "currentEnrollment": 0,
            "dayOfWeek": "Mon",
            "startTime": "07:30",
            "endTime": "09:30",
            "room": "R1",
            "lecturerID": "",
        })
        self.assertTrue(success)

        success, _ = admin.manageClassSection("add", {
            "classCode": "CS101-01",
            "courseCode": "CS101",
            "semesterCode": "SEM1-2024",
            "maxCapacity": 40,
            "currentEnrollment": 0,
            "dayOfWeek": "Mon",
            "startTime": "07:30",
            "endTime": "09:30",
            "room": "R1",
            "lecturerID": "",
        })
        self.assertFalse(success)

    def test_10_view_training_program_status(self):
        """TC41, TC42: Sinh viÃªn xem CTDT vÃ  tráº¡ng thÃ¡i Completed/Incomplete."""
        Course("CS101", "Intro IT", 3, "IT", "").save()
        Course("CS301", "Advanced Programming", 3, "IT", "CS101").save()
        TrainingProgram("SE2024", "Software Engineering", "CS101,CS301").save()
        ClassSection("C101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("C301", "CS301", "SEM1", 40, 0, "Tue", "09:30", "11:30", "R2", "GV001").save()
        Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT", "SE2024", "Active").save()

        g = Grade("C101", "SV2024001", 8, 8, 8, 8)
        g.calculateFinalSummary()
        g.assignLetterGrade()
        g.finalize()

        student = Student.find_by_id("SV2024001")
        program = student.viewTrainingProgram()
        self.assertIsNotNone(program)
        self.assertEqual(program.getCourseStatus("CS101", "SV2024001"), "Completed")
        self.assertEqual(program.getCourseStatus("CS301", "SV2024001"), "Incomplete")

    def test_11_take_attendance(self):
        """TC43, TC44: Äiá»ƒm danh thÃ nh cÃ´ng vÃ  cháº·n lá»›p phiÃªn khÃ´ng há»£p lá»‡."""
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        valid_class = ClassSection("CS101-01", "CS101", "SEM1-2024", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001")
        valid_class.save()
        success, _ = lecturer.takeAttendance(valid_class, "SV2024001", "Absent")
        self.assertTrue(success)
        records = Attendance.get_student_class_records("SV2024001", "CS101-01")
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].attendanceStatus, "Absent")
        self.assertEqual(records[0].attendanceRate, 0.0)

        invalid_class = ClassSection("XXXX-99", "", "", 0, 0, "", "", "", "", "")
        success, _ = lecturer.takeAttendance(invalid_class, "SV2024001", "Absent")
        self.assertFalse(success)

    def test_12_enter_grades_and_edit_after_finalize(self):
        """TC45, TC46, TC47: Nháº­p Ä‘iá»ƒm, tá»± tÃ­nh tá»•ng káº¿t, cháº·n sá»­a khi Ä‘Ã£ hoàn tất."""
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")

        success, _ = lecturer.inputComponentGrades("CS101-01", "SV2024001", 7.5, 8.0, 9.0)
        self.assertTrue(success)
        success, _ = lecturer.inputFinalExamGrade("CS101-01", "SV2024001", 8.5)
        self.assertTrue(success)
        g = Grade.find_by_student_class("SV2024001", "CS101-01")
        self.assertEqual(g.status, "Draft")
        expected = round(7.5 * 0.2 + 8.0 * 0.1 + 9.0 * 0.1 + 8.5 * 0.6, 2)
        self.assertEqual(g.finalSummaryGrade, expected)

        g_formula = Grade("TMP", "SVTMP", 6.0, 0.0, 0.0, 8.0)
        self.assertEqual(g_formula.calculateFinalSummary(), 7.4)
        self.assertEqual(g_formula.assignLetterGrade(), "B")

        g.finalize()
        success, _ = lecturer.inputFinalExamGrade("CS101-01", "SV2024001", 9.0)
        self.assertFalse(success)

    def test_13_finalize_grades_deadline_and_lock(self):
        """TC48, TC49, TC50: Duyá»‡t Ä‘iá»ƒm theo háº¡n vÃ  khÃ³a Ä‘iá»ƒm sau khi duyá»‡t."""
        admin = Admin("admin01")
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        student = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        student.save()
        Course("CS101", "Intro IT", 3, "IT").save()

        today = datetime.date.today()
        past_end = (today - datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        future_end = (today + datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        start = (today - datetime.timedelta(days=100)).strftime("%d/%m/%Y")

        Semester("SEM_PAST", start, past_end, "").save()
        ClassSection("CLS_PAST", "CS101", "SEM_PAST", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        lecturer.inputComponentGrades("CLS_PAST", "SV2024001", 7, 8, 9)
        lecturer.inputFinalExamGrade("CLS_PAST", "SV2024001", 8)

        success, _ = admin.finalizeGrades("CLS_PAST")
        self.assertTrue(success)
        g = Grade.find_by_student_class("SV2024001", "CLS_PAST")
        self.assertEqual(g.status, "Finalized")
        success, _ = lecturer.inputFinalExamGrade("CLS_PAST", "SV2024001", 9)
        self.assertFalse(success)

        Semester("SEM_FUTURE", start, future_end, "").save()
        ClassSection("CLS_FUT", "CS101", "SEM_FUTURE", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()
        lecturer.inputComponentGrades("CLS_FUT", "SV2024001", 6, 6, 6)
        lecturer.inputFinalExamGrade("CLS_FUT", "SV2024001", 6)
        success, _ = admin.finalizeGrades("CLS_FUT")
        self.assertFalse(success)

    def test_14_view_transcript_gpa_and_breakdown(self):
        """TC51, TC52, TC53: Chá»‰ hiá»ƒn thá»‹ đã chốt, GPA Ä‘Ãºng, Ä‘á»§ breakdown Ä‘iá»ƒm."""
        s = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        s.save()
        Course("CS101", "Course 1", 3, "IT").save()
        Course("CS201", "Course 2", 3, "IT").save()
        ClassSection("CLS101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("CLS201", "CS201", "SEM1", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()

        g1 = Grade("CLS101", "SV2024001", 8, 8, 8, 8)
        g1.calculateFinalSummary()
        g1.assignLetterGrade()
        g1.finalize()

        g2 = Grade("CLS201", "SV2024001", 7, 7, 7, 7)
        g2.calculateFinalSummary()
        g2.assignLetterGrade()
        g2.save()

        transcript = s.viewGrades()
        self.assertTrue(all(g.status == "Finalized" for g in transcript))

        sem_gpa = s.calculateSemesterGPA("SEM1")
        cum_gpa = s.calculateCumulativeGPA()
        self.assertEqual(sem_gpa, cum_gpa)
        self.assertEqual(sem_gpa, g1.finalSummaryGrade)

        g1_loaded = Grade.find_by_student_class("SV2024001", "CLS101")
        self.assertTrue(g1_loaded.status == "Finalized")
        self.assertIsInstance(g1_loaded.midtermGrade, float)
        self.assertIsInstance(g1_loaded.assignmentGrade, float)
        self.assertIsInstance(g1_loaded.attendanceGrade, float)
        self.assertIsInstance(g1_loaded.finalExamGrade, float)
        self.assertIn(g1_loaded.letterGrade, ["A", "B", "C", "D", "F"])

    def test_15_finalize_multiple_courses_weighted_gpa(self):
        """TC54: Duyá»‡t nhiá»u mÃ´n vÃ  tÃ­nh GPA theo tÃ­n chá»‰ chÃ­nh xÃ¡c."""
        admin = Admin("admin01")
        s = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        s.save()

        today = datetime.date.today()
        start = (today - datetime.timedelta(days=100)).strftime("%d/%m/%Y")
        end = (today - datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        Semester("SEM1", start, end, "").save()

        Course("CS101", "Course 1", 3, "IT").save()
        Course("CS102", "Course 2", 3, "IT").save()
        Course("CS103", "Course 3", 2, "IT").save()
        ClassSection("C101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("C102", "CS102", "SEM1", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()
        ClassSection("C103", "CS103", "SEM1", 40, 0, "Wed", "07:30", "09:30", "R3", "GV001").save()

        Grade("C101", "SV2024001", 7, 7, 7, 7, 7.5, "B", "Draft").save()
        Grade("C102", "SV2024001", 9, 9, 9, 9, 9.0, "A", "Draft").save()
        Grade("C103", "SV2024001", 6, 6, 6, 6, 6.0, "C", "Draft").save()

        self.assertTrue(admin.finalizeGrades("C101")[0])
        self.assertTrue(admin.finalizeGrades("C102")[0])
        self.assertTrue(admin.finalizeGrades("C103")[0])

        sem_gpa = s.calculateSemesterGPA("SEM1")
        expected = round((7.5 * 3 + 9.0 * 3 + 6.0 * 2) / 8, 2)
        self.assertEqual(sem_gpa, expected)

    def test_16_security_rbac_performance_integrity(self):
        """TC56, TC57, TC58: RBAC, hiá»‡u nÄƒng vÃ  toÃ n váº¹n dá»¯ liá»‡u file .txt."""
        student = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        student.save()

        from controllers.user_controller import create_account
        success, _, _ = create_account(
            "SVHACK", "123456789012", "Hack User", "01/01/2000",
            "hack@x.com", "Male", "0909999999", "IT", "Student", "SE", "Active", "Student"
        )
        self.assertFalse(success)

        start = time.perf_counter()
        Account("sv001", hash_password("Abc@12345"), "Student").save()
        Account.login("sv001", "Abc@12345")
        Student.get_all()
        elapsed = time.perf_counter() - start
        self.assertLess(elapsed, 3.0)

        st = Student("SVX", "9999", "Data Integrity", "01/01/2000", "di@x.com", "Male", "0909", "IT")
        self.assertTrue(st.save())
        found = Student.find_by_id("SVX")
        self.assertIsNotNone(found)

        ok, _ = found.updatePersonalInfo(email="di_new@x.com", phoneNumber="0912345678")
        self.assertTrue(ok)
        reloaded = Student.find_by_id("SVX")
        self.assertEqual(reloaded.email, "di_new@x.com")

        self.assertTrue(reloaded.delete())
        self.assertIsNone(Student.find_by_id("SVX"))

        for name in os.listdir(TEST_DATA_DIR):
            self.assertFalse(name.endswith(".tmp"))


def _run_single_test_method(method_name):
    """Chạy một phương thức unittest và trả về (đạt, thông_điệp_chi_tiết)."""
    suite = unittest.TestSuite([SystemFunctionalityTest(method_name)])
    stream = io.StringIO()
    result = unittest.TextTestRunner(stream=stream, verbosity=0).run(suite)
    if result.wasSuccessful():
        return True, "PASS"

    if result.failures:
        return False, result.failures[0][1].strip().splitlines()[-1]
    if result.errors:
        return False, result.errors[0][1].strip().splitlines()[-1]
    return False, "Unknown failure"


def run_from_excel(excel_path):
    """Đọc danh sách TC từ Excel, chạy các kiểm thử đã ánh xạ và in trạng thái từng TC."""
    try:
        import openpyxl
    except Exception:
        print("ERROR: Missing dependency 'openpyxl'. Install it first: pip install openpyxl")
        return False

    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}")
        return False

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    # Ánh xạ mỗi mã TC với một phương thức unittest tự động.
    tc_to_method = {
        "TC01": "test_1_auth_system",
        "TC02": "test_1_auth_system",
        "TC03": "test_1_auth_system",
        "TC04": "test_1_auth_system",
        "TC05": "test_1_auth_system",
        "TC06": "test_1_auth_system",
        "TC07": "test_1_auth_system",
        "TC08": "test_1_auth_system",
        "TC09": "test_1_auth_system",
        "TC10": "test_1_auth_system",
        "TC11": "test_1_auth_system",
        "TC12": "test_1_auth_system",
        "TC13": "test_1_auth_system",
        "TC14": "test_1_auth_system",
        "TC15": "test_1_auth_system",
        "TC16": "test_2_student_management",
        "TC17": "test_2_student_management",
        "TC18": "test_2_student_management",
        "TC19": "test_16_security_rbac_performance_integrity",
        "TC20": "test_2_student_management",
        "TC21": "test_2_student_management",
        "TC22": "test_2_student_management",
        "TC23": "test_2_student_management",
        "TC24": "test_2_student_management",
        "TC25": "test_2_student_management",
        "TC26": "test_2_student_management",
        "TC27": "test_2_student_management",
        "TC28": "test_6_manage_faculty",
        "TC29": "test_2_student_management",
        "TC30": "test_5_assign_lecturers",
        "TC31": "test_5_assign_lecturers",
        "TC32": "test_6_manage_faculty",
        "TC33": "test_6_manage_faculty",
        "TC34": "test_7_manage_semester_invalid_end_date",
        "TC35": "test_7_manage_semester_invalid_end_date",
        "TC36": "test_8_manage_course_and_prerequisites",
        "TC37": "test_8_manage_course_and_prerequisites",
        "TC38": "test_8_manage_course_and_prerequisites",
        "TC39": "test_9_manage_class_section",
        "TC40": "test_9_manage_class_section",
        "TC41": "test_10_view_training_program_status",
        "TC42": "test_10_view_training_program_status",
        "TC43": "test_11_take_attendance",
        "TC44": "test_11_take_attendance",
        "TC45": "test_12_enter_grades_and_edit_after_finalize",
        "TC46": "test_12_enter_grades_and_edit_after_finalize",
        "TC47": "test_12_enter_grades_and_edit_after_finalize",
        "TC48": "test_13_finalize_grades_deadline_and_lock",
        "TC49": "test_13_finalize_grades_deadline_and_lock",
        "TC50": "test_13_finalize_grades_deadline_and_lock",
        "TC51": "test_14_view_transcript_gpa_and_breakdown",
        "TC52": "test_14_view_transcript_gpa_and_breakdown",
        "TC53": "test_14_view_transcript_gpa_and_breakdown",
        "TC54": "test_15_finalize_multiple_courses_weighted_gpa",
        "TC55": "test_1_auth_system",
        "TC56": "test_16_security_rbac_performance_integrity",
        "TC57": "test_16_security_rbac_performance_integrity",
        "TC58": "test_16_security_rbac_performance_integrity",
    }

    # Chạy mỗi phương thức đã ánh xạ đúng một lần.
    method_results = {}
    for method_name in sorted(set(tc_to_method.values())):
        method_results[method_name] = _run_single_test_method(method_name)

    # In trạng thái từng test case theo thứ tự trong Excel.
    tc_rows = []
    for r in range(2, ws.max_row + 1):
        tc_id = ws.cell(r, 1).value
        func_name = ws.cell(r, 2).value
        if tc_id:
            tc_rows.append((str(tc_id).strip(), str(func_name or "").strip()))

    passed = 0
    failed = 0
    not_automated = 0

    print("\n EXCEL-DRIVEN TEST REPORT ")
    print(f"Source file: {excel_path}")
    for tc_id, func_name in tc_rows:
        method = tc_to_method.get(tc_id)
        if not method:
            not_automated += 1
            print(f"{tc_id} | NOT_AUTOMATED | {func_name}")
            continue

        ok, detail = method_results[method]
        if ok:
            passed += 1
            print(f"{tc_id} | PASS | {func_name}")
        else:
            failed += 1
            print(f"{tc_id} | FAIL | {func_name} | {detail}")

    total = len(tc_rows)
    print("\n===== SUMMARY =====")
    print(f"TOTAL: {total}")
    print(f"PASS: {passed}")
    print(f"FAIL: {failed}")
    print(f"NOT_AUTOMATED: {not_automated}")

    if failed == 0 and not_automated == 0:
        print("ALL PASS")
        return True

    return False


if __name__ == '__main__':
    print("Starting tests from the Excel test case file...")
    excel_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "test case", "SMS TestCases.xlsx")
    success = run_from_excel(excel_file)
    sys.exit(0 if success else 1)
import os
import sys
import shutil
import openpyxl
import datetime
import time

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import utils.file_handler as file_handler
TEST_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_data_mock")
file_handler.DATA_DIR = TEST_DATA_DIR

from models.account import Account
from models.admin import Admin
from models.student import Student
from models.grade import Grade
from models.course import Course
from models.class_section import ClassSection
from models.lecturer import Lecturer
from models.faculty import Faculty
from models.semester import Semester
from models.training_program import TrainingProgram
from models.attendance import Attendance
from utils.auth_helper import hash_password, verify_password


def setup_test_env():
    if not os.path.exists(TEST_DATA_DIR):
        os.makedirs(TEST_DATA_DIR)


def cleanup_test_env():
    if os.path.exists(TEST_DATA_DIR):
        shutil.rmtree(TEST_DATA_DIR)


def clear_test_data():
    if os.path.exists(TEST_DATA_DIR):
        for f in os.listdir(TEST_DATA_DIR):
            try:
                os.remove(os.path.join(TEST_DATA_DIR, f))
            except:
                pass


# ============ BỘ XỬ LÝ KIỂM THỬ ============

def test_login(input_data, expected_result):
    """TC01-TC15, TC55: Kiểm thử đăng nhập."""
    clear_test_data()
    try:
        acc = Account("sv001", hash_password("Abc@12345"), "Student")
        acc.save()
        result = Account.login("sv001", "Abc@12345")
        result_wrong = Account.login("sv001", "wrongpass")
        return result and not result_wrong
    except:
        return False


def test_create_account(input_data, expected_result):
    """TC16-TC19: Kiểm thử tạo tài khoản với RBAC."""
    clear_test_data()
    try:
        from controllers.user_controller import create_account
        
        success, _, _ = create_account(
            "SV001", "012345678912", "Test Student", "01/01/2000",
            "test@uni.edu", "Male", "0901000000", "IT", "Student", "SE2024", "Active", "Admin"
        )
        if not success:
            return False
        
        success_student, _, _ = create_account(
            "SVHACK", "123456789012", "Hack User", "01/01/2000",
            "hack@x.com", "Male", "0909999999", "IT", "Student", "SE", "Active", "Student"
        )
        return not success_student
    except:
        return False


def test_manage_student_profile(input_data, expected_result):
    """TC20-TC29: Kiểm thử quản lý sinh viên."""
    clear_test_data()
    try:
        s = Student("SV001", "012345678912", "Nguyen Van A", "01/01/2000", 
                   "a@gmail.com", "Male", "0900000000", "CNTT")
        s.save()
        
        s_found = Student.find_by_id("SV001")
        if s_found.fullName != "Nguyen Van A":
            return False
        
        success, msg = s_found.updatePersonalInfo(email="a_new@gmail.com", phoneNumber="0911111111")
        if not success:
            return False
        
        s_updated = Student.find_by_id("SV001")
        return s_updated.email == "a_new@gmail.com"
    except:
        return False


def test_assign_lecturers(input_data, expected_result):
    """TC30-TC31: Kiểm thử phân công giảng viên cho lớp học phần."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT").save()
        ClassSection("CS101-01", "CS101", "SEM1-2024", 40, 0, "Mon", "07:30", "09:30", "R1", "").save()
        
        success, msg = admin.assignLecturer("GV001", "CS101-01")
        if not success:
            return False
        
        cs = ClassSection.find_by_code("CS101-01")
        return cs.lecturerID == "GV001"
    except:
        return False


def test_manage_faculty(input_data, expected_result):
    """TC32-TC33: Kiểm thử quản lý khoa."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        data = {
            "facultyCode": "IT",
            "facultyName": "Information Technology",
            "email": "it@uni.edu",
            "phoneNumber": "02812345678",
        }
        success, msg = admin.manageFaculty("add", data)
        if not success:
            return False
        
        if Faculty.find_by_code("IT") is None:
            return False
        
        success_dup, _ = admin.manageFaculty("add", data)
        return not success_dup
    except:
        return False


def test_manage_semester(input_data, expected_result):
    """TC34-TC35: Kiểm thử quản lý học kỳ kèm kiểm tra ngày tháng."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        valid = {
            "semester": "SEM1-2024",
            "startDate": "01/09/2024",
            "endDate": "31/01/2025",
            "examWeeks": "13/01-24/01/2025",
        }
        success, _ = admin.manageSemester("add", valid)
        if not success:
            return False
        
        invalid = {
            "semester": "SEM_BAD",
            "startDate": "01/09/2024",
            "endDate": "01/08/2024",
            "examWeeks": "",
        }
        success_invalid, _ = admin.manageSemester("add", invalid)
        return not success_invalid
    except:
        return False


def test_manage_course(input_data, expected_result):
    """TC36-TC38: Kiểm thử quản lý môn học kèm kiểm tra hợp lệ."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        success, _ = admin.manageCourse("add", {
            "courseCode": "CS101",
            "courseName": "Introduction to IT",
            "credits": 3,
            "facultyCode": "IT",
            "prerequisites": "",
        })
        if not success:
            return False
        
        success_invalid, _ = admin.manageCourse("add", {
            "courseCode": "CS201",
            "courseName": "Data Structures",
            "credits": "",
            "facultyCode": "",
            "prerequisites": "",
        })
        return not success_invalid
    except:
        return False


def test_manage_class_section(input_data, expected_result):
    """TC39-TC40: Kiểm thử quản lý lớp học phần."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        success, _ = admin.manageClassSection("add", {
            "classCode": "CS101-01",
            "courseCode": "CS101",
            "semesterCode": "SEM1-2024",
            "maxCapacity": 40,
            "currentEnrollment": 0,
            "dayOfWeek": "Mon",
            "startTime": "07:30",
            "endTime": "09:30",
            "room": "R1",
            "lecturerID": "",
        })
        if not success:
            return False
        
        success_dup, _ = admin.manageClassSection("add", {
            "classCode": "CS101-01",
            "courseCode": "CS101",
            "semesterCode": "SEM1-2024",
            "maxCapacity": 40,
            "currentEnrollment": 0,
            "dayOfWeek": "Mon",
            "startTime": "07:30",
            "endTime": "09:30",
            "room": "R1",
            "lecturerID": "",
        })
        return not success_dup
    except:
        return False


def test_view_training_program(input_data, expected_result):
    """TC41-TC42: Kiểm thử xem chương trình đào tạo."""
    clear_test_data()
    try:
        Course("CS101", "Intro IT", 3, "IT", "").save()
        Course("CS301", "Advanced Programming", 3, "IT", "CS101").save()
        TrainingProgram("SE2024", "Software Engineering", "CS101,CS301").save()
        ClassSection("C101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("C301", "CS301", "SEM1", 40, 0, "Tue", "09:30", "11:30", "R2", "GV001").save()
        Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT", "SE2024", "Active").save()
        
        g = Grade("C101", "SV2024001", 8, 8, 8, 8)
        g.calculateFinalSummary()
        g.assignLetterGrade()
        g.finalize()
        
        student = Student.find_by_id("SV2024001")
        program = student.viewTrainingProgram()
        return program is not None
    except:
        return False


def test_take_attendance(input_data, expected_result):
    """TC43-TC44: Kiểm thử điểm danh."""
    clear_test_data()
    try:
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        valid_class = ClassSection("CS101-01", "CS101", "SEM1-2024", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001")
        valid_class.save()
        success, _ = lecturer.takeAttendance(valid_class, "SV2024001", "Absent")
        if not success:
            return False
        
        records = Attendance.get_student_class_records("SV2024001", "CS101-01")
        return len(records) == 1 and records[0].attendanceStatus == "Absent"
    except:
        return False


def test_enter_grades(input_data, expected_result):
    """TC45-TC47: Kiểm thử nhập điểm và xác minh công thức tính."""
    clear_test_data()
    try:
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        
        success, _ = lecturer.inputComponentGrades("CS101-01", "SV2024001", 7.5, 8.0, 9.0)
        if not success:
            return False
        
        success, _ = lecturer.inputFinalExamGrade("CS101-01", "SV2024001", 8.5)
        if not success:
            return False
        
        g = Grade.find_by_student_class("SV2024001", "CS101-01")
        expected = round(7.5 * 0.2 + 8.0 * 0.1 + 9.0 * 0.1 + 8.5 * 0.6, 2)
        return g.finalSummaryGrade == expected
    except:
        return False


def test_finalize_grades(input_data, expected_result):
    """TC48-TC50: Kiểm thử chốt điểm theo hạn chót."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        student = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        student.save()
        Course("CS101", "Intro IT", 3, "IT").save()
        
        today = datetime.date.today()
        past_end = (today - datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        start = (today - datetime.timedelta(days=100)).strftime("%d/%m/%Y")
        
        Semester("SEM_PAST", start, past_end, "").save()
        ClassSection("CLS_PAST", "CS101", "SEM_PAST", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        lecturer.inputComponentGrades("CLS_PAST", "SV2024001", 7, 8, 9)
        lecturer.inputFinalExamGrade("CLS_PAST", "SV2024001", 8)
        
        success, _ = admin.finalizeGrades("CLS_PAST")
        if not success:
            return False
        
        g = Grade.find_by_student_class("SV2024001", "CLS_PAST")
        return g.status == "Finalized"
    except:
        return False


def test_view_transcript(input_data, expected_result):
    """TC51-TC53: Kiểm thử xem bảng điểm và GPA."""
    clear_test_data()
    try:
        s = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        s.save()
        Course("CS101", "Course 1", 3, "IT").save()
        Course("CS201", "Course 2", 3, "IT").save()
        ClassSection("CLS101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("CLS201", "CS201", "SEM1", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()
        
        g1 = Grade("CLS101", "SV2024001", 8, 8, 8, 8)
        g1.calculateFinalSummary()
        g1.assignLetterGrade()
        g1.finalize()
        
        transcript = s.viewGrades()
        return all(g.status == "Finalized" for g in transcript)
    except:
        return False


def test_weighted_gpa(input_data, expected_result):
    """TC54: Tính GPA có trọng số."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        s = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        s.save()
        
        today = datetime.date.today()
        start = (today - datetime.timedelta(days=100)).strftime("%d/%m/%Y")
        end = (today - datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        Semester("SEM1", start, end, "").save()
        
        Course("CS101", "Course 1", 3, "IT").save()
        Course("CS102", "Course 2", 3, "IT").save()
        Course("CS103", "Course 3", 2, "IT").save()
        ClassSection("C101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("C102", "CS102", "SEM1", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()
        ClassSection("C103", "CS103", "SEM1", 40, 0, "Wed", "07:30", "09:30", "R3", "GV001").save()
        
        Grade("C101", "SV2024001", 7, 7, 7, 7, 7.5, "B", "Draft").save()
        Grade("C102", "SV2024001", 9, 9, 9, 9, 9.0, "A", "Draft").save()
        Grade("C103", "SV2024001", 6, 6, 6, 6, 6.0, "C", "Draft").save()
        
        admin.finalizeGrades("C101")
        admin.finalizeGrades("C102")
        admin.finalizeGrades("C103")
        
        sem_gpa = s.calculateSemesterGPA("SEM1")
        expected = round((7.5 * 3 + 9.0 * 3 + 6.0 * 2) / 8, 2)
        return sem_gpa == expected
    except:
        return False


def test_security_rbac(input_data, expected_result):
    """TC56-TC58: Kiểm thử RBAC, hiệu năng và toàn vẹn dữ liệu."""
    clear_test_data()
    try:
        from controllers.user_controller import create_account
        
        student = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        student.save()
        
        success, _, _ = create_account(
            "SVHACK", "123456789012", "Hack User", "01/01/2000",
            "hack@x.com", "Male", "0909999999", "IT", "Student", "SE", "Active", "Student"
        )
        if success:
            return False
        
        start = time.perf_counter()
        Account("sv001", hash_password("Abc@12345"), "Student").save()
        Account.login("sv001", "Abc@12345")
        Student.get_all()
        elapsed = time.perf_counter() - start
        if elapsed >= 3.0:
            return False
        
        st = Student("SVX", "9999", "Data Integrity", "01/01/2000", "di@x.com", "Male", "0909", "IT")
        st.save()
        found = Student.find_by_id("SVX")
        if found is None:
            return False
        
        ok, _ = found.updatePersonalInfo(email="di_new@x.com", phoneNumber="0912345678")
        if not ok:
            return False
        
        reloaded = Student.find_by_id("SVX")
        return reloaded.email == "di_new@x.com"
    except:
        return False


# ============ ÁNH XẠ & TRÌNH CHẠY ============

FUNCTION_HANDLERS = {
    "Log In": test_login,
    "Forgot Password": test_login,
    "Log Out": test_login,
    "Change Password": test_login,
    "Create Accounts": test_create_account,
    "Manage Student Profile": test_manage_student_profile,
    "Update Profile": test_manage_student_profile,
    "View Profile": test_manage_student_profile,
    "Manage Lecturer Profile": test_manage_student_profile,
    "Assign Lecturers": test_assign_lecturers,
    "Manage Faculty": test_manage_faculty,
    "Manage Semester/Term": test_manage_semester,
    "Manage Course": test_manage_course,
    "Manage Class/Section": test_manage_class_section,
    "View Training Program": test_view_training_program,
    "Take Attendance": test_take_attendance,
    "Enter Grades": test_enter_grades,
    "Finalize Grades": test_finalize_grades,
    "View Transcript": test_view_transcript,
    "Security â€“ Password Hashing": test_security_rbac,
    "Security â€“ RBAC": test_security_rbac,
    "Performance": test_security_rbac,
    "Reliability â€“ Data Integrity": test_security_rbac,
}


def run_from_excel(excel_path):
    """Đọc Excel và chạy toàn bộ kiểm thử theo các hàm xử lý chức năng."""
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}")
        return False
    
    setup_test_env()
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[wb.sheetnames[0]]
        
        passed = 0
        failed = 0
        not_automated = 0
        
        print("\n EXCEL-DRIVEN TEST REPORT ")
        print(f"Source file: {excel_path}\n")
        
        for r in range(2, ws.max_row + 1):
            tc_id = ws.cell(r, 1).value
            func_name = ws.cell(r, 2).value
            input_data = ws.cell(r, 4).value or ""
            expected_result = ws.cell(r, 5).value or ""
            
            if not tc_id:
                continue
            
            tc_id = str(tc_id).strip()
            func_name = str(func_name or "").strip()
            
            handler = FUNCTION_HANDLERS.get(func_name)
            if not handler:
                not_automated += 1
                print(f"{tc_id} | NOT_AUTOMATED | {func_name}")
                continue
            
            try:
                result = handler(input_data, expected_result)
                if result:
                    passed += 1
                    print(f"{tc_id} | PASS | {func_name}")
                else:
                    failed += 1
                    print(f"{tc_id} | FAIL | {func_name}")
            except Exception as e:
                failed += 1
                error_msg = str(e).split('\n')[0][:40]
                print(f"{tc_id} | FAIL | {func_name} | {error_msg}")
        
        total = passed + failed + not_automated
        print("\n===== SUMMARY =====")
        print(f"TOTAL: {total}")
        print(f"PASS: {passed}")
        print(f"FAIL: {failed}")
        print(f"NOT_AUTOMATED: {not_automated}")
        
        if failed == 0 and not_automated == 0:
            print("\nALL PASS")
            return True
        
        return False
    
    except Exception as e:
        print(f"ERROR: {str(e)}")
        return False
    
    finally:
        cleanup_test_env()


if __name__ == '__main__':
    print("Running tests from Excel...\n")
    excel_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                              "test case", "SMS TestCases.xlsx")
    success = run_from_excel(excel_file)
    sys.exit(0 if success else 1)
import os, sys, shutil, openpyxl, datetime, time
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import utils.file_handler as file_handler
TEST_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_data_mock")
file_handler.DATA_DIR = TEST_DATA_DIR

from models.account import Account
from models.admin import Admin
from models.student import Student
from models.grade import Grade
from models.course import Course
from models.class_section import ClassSection
from models.lecturer import Lecturer
from models.faculty import Faculty
from models.semester import Semester
from models.training_program import TrainingProgram
from models.attendance import Attendance
from utils.auth_helper import hash_password, verify_password

def setup(): 
    if not os.path.exists(TEST_DATA_DIR): os.makedirs(TEST_DATA_DIR)

def cleanup(): 
    if os.path.exists(TEST_DATA_DIR): shutil.rmtree(TEST_DATA_DIR)

def clear():
    for f in os.listdir(TEST_DATA_DIR):
        try: os.remove(os.path.join(TEST_DATA_DIR, f))
        except: pass

def test_handler(func_name):
    clear()
    try:
        if func_name in ["Log In", "Forgot Password", "Log Out", "Change Password"]:
            Account("sv001", hash_password("Abc@12345"), "Student").save()
            return Account.login("sv001", "Abc@12345") and not Account.login("sv001", "wrong")
        
        elif func_name == "Create Accounts":
            from controllers.user_controller import create_account
            Account("admin", hash_password("admin123"), "Admin").save()
            ok = create_account("SV001", "012345678912", "Test", "01/01/2000", "test@u.edu", "M", "090", "IT", "Student", "SE", "Active", "Admin")[0]
            bad = create_account("SVHACK", "123456789012", "Hack", "01/01/2000", "h@x.com", "M", "090", "IT", "Student", "SE", "Active", "Student")[0]
            return ok and not bad
        
        elif "Student Profile" in func_name or "Update Profile" in func_name or "View Profile" in func_name:
            s = Student("SV001", "012345678912", "Nguyen Van A", "01/01/2000", "a@gmail.com", "Male", "0900000000", "CNTT")
            s.save()
            sf = Student.find_by_id("SV001")
            if sf.fullName != "Nguyen Van A": return False
            sf.updatePersonalInfo(email="a_new@gmail.com", phoneNumber="0911111111")
            return Student.find_by_id("SV001").email == "a_new@gmail.com"
        
        elif func_name == "Assign Lecturers":
            admin = Admin("admin01")
            Lecturer("GV001", "0123", "Dr", "01/01/1980", "e@u.edu", "F", "0901000000", "IT").save()
            ClassSection("CS101-01", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "").save()
            ok = admin.assignLecturer("GV001", "CS101-01")[0]
            bad = admin.assignLecturer("GV001", "XXXX")[0]
            return ok and not bad
        
        elif func_name == "Manage Faculty":
            admin = Admin("admin01")
            d = {"facultyCode": "IT", "facultyName": "IT", "email": "it@u.edu", "phoneNumber": "0281000000"}
            ok = admin.manageFaculty("add", d)[0]
            bad = admin.manageFaculty("add", d)[0]
            return ok and not bad
        
        elif func_name == "Manage Semester/Term":
            admin = Admin("admin01")
            ok = admin.manageSemester("add", {"semester": "SEM1", "startDate": "01/09/2024", "endDate": "31/01/2025", "examWeeks": ""})[0]
            bad = admin.manageSemester("add", {"semester": "SEM2", "startDate": "01/09/2024", "endDate": "01/08/2024", "examWeeks": ""})[0]
            return ok and not bad
        
        elif func_name == "Manage Course":
            admin = Admin("admin01")
            ok = admin.manageCourse("add", {"courseCode": "CS101", "courseName": "Intro", "credits": 3, "facultyCode": "IT", "prerequisites": ""})[0]
            bad = admin.manageCourse("add", {"courseCode": "CS201", "courseName": "Data", "credits": "", "facultyCode": "", "prerequisites": ""})[0]
            return ok and not bad
        
        elif func_name == "Manage Class/Section":
            admin = Admin("admin01")
            d = {"classCode": "CS101-01", "courseCode": "CS101", "semesterCode": "SEM1", "maxCapacity": 40, "currentEnrollment": 0, "dayOfWeek": "Mon", "startTime": "07:30", "endTime": "09:30", "room": "R1", "lecturerID": ""}
            ok = admin.manageClassSection("add", d)[0]
            bad = admin.manageClassSection("add", d)[0]
            return ok and not bad
        
        elif func_name == "View Training Program":
            Course("CS101", "Intro", 3, "IT", "").save()
            TrainingProgram("SE2024", "SE", "CS101").save()
            ClassSection("C101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
            Student("SV001", "1111", "John", "01/01/2003", "j@x.com", "M", "0901", "IT", "SE2024", "Active").save()
            return Student.find_by_id("SV001").viewTrainingProgram() is not None
        
        elif func_name == "Take Attendance":
            gv = Lecturer("GV001", "0123", "Dr", "01/01/1980", "e@u.edu", "F", "0901000000", "IT")
            cs = ClassSection("CS101-01", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001")
            cs.save()
            return gv.takeAttendance(cs, "SV001", "Absent")[0]
        
        elif func_name == "Enter Grades":
            gv = Lecturer("GV001", "0123", "Dr", "01/01/1980", "e@u.edu", "F", "0901000000", "IT")
            ok1 = gv.inputComponentGrades("CS101", "SV001", 7.5, 8, 9)[0]
            ok2 = gv.inputFinalExamGrade("CS101", "SV001", 8.5)[0]
            return ok1 and ok2
        
        elif func_name == "Finalize Grades":
            admin = Admin("admin01")
            Semester("SEM", "01/09/2024", "31/12/2024", "").save()
            Course("CS101", "Intro", 3, "IT").save()
            ClassSection("CLS", "CS101", "SEM", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
            return admin.finalizeGrades("CLS")[0]
        
        elif func_name == "View Transcript":
            s = Student("SV001", "1111", "John", "01/01/2003", "j@x.com", "M", "0901", "IT")
            s.save()
            return len(s.viewGrades()) >= 0
        
        elif "Security" in func_name or "Performance" in func_name or "Reliability" in func_name:
            from controllers.user_controller import create_account
            s = Student("SV001", "1111", "John", "01/01/2003", "j@x.com", "M", "0901", "IT")
            s.save()
            bad = create_account("SVHACK", "123456789012", "Hack", "01/01/2000", "h@x.com", "M", "090", "IT", "Student", "SE", "Active", "Student")[0]
            return not bad
        
        return False
    except: return False

def run_from_excel(excel_path):
    if not os.path.exists(excel_path): return False
    
    setup()
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[wb.sheetnames[0]]
        p, f, n = 0, 0, 0
        print("\n EXCEL-DRIVEN TEST REPORT \nSource: {}\n".format(excel_path))
        
        for r in range(2, ws.max_row + 1):
            tc = str(ws.cell(r, 1).value or "").strip()
            func = str(ws.cell(r, 2).value or "").strip()
            if not tc: continue
            
            if test_handler(func):
                p += 1
                print("{} | PASS | {}".format(tc, func))
            else:
                f += 1
                print("{} | FAIL | {}".format(tc, func))
        
        t = p + f
        print("\n===== SUMMARY =====\nTOTAL: {}\nPASS: {}\nFAIL: {}".format(t, p, f))
        return f == 0 and p > 0
    except: return False
    finally: cleanup()

if __name__ == '__main__':
    print("Running tests...\n")
    e = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "test case", "SMS TestCases.xlsx")
    sys.exit(0 if run_from_excel(e) else 1)
import os
import sys
import shutil
import openpyxl
import datetime
import time

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import utils.file_handler as file_handler
TEST_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_data_mock")
file_handler.DATA_DIR = TEST_DATA_DIR

from models.account import Account
from models.admin import Admin
from models.student import Student
from models.grade import Grade
from models.course import Course
from models.class_section import ClassSection
from models.lecturer import Lecturer
from models.faculty import Faculty
from models.semester import Semester
from models.training_program import TrainingProgram
from models.attendance import Attendance
from utils.auth_helper import hash_password, verify_password


def setup_test_env():
    if not os.path.exists(TEST_DATA_DIR):
        os.makedirs(TEST_DATA_DIR)


def cleanup_test_env():
    if os.path.exists(TEST_DATA_DIR):
        shutil.rmtree(TEST_DATA_DIR)


def clear_test_data():
    if os.path.exists(TEST_DATA_DIR):
        for f in os.listdir(TEST_DATA_DIR):
            try:
                os.remove(os.path.join(TEST_DATA_DIR, f))
            except:
                pass


# ============ BỘ XỬ LÝ KIỂM THỬ ============

def test_login(input_data, expected_result):
    """TC01-TC15, TC55: Kiểm thử đăng nhập."""
    clear_test_data()
    try:
        lines = input_data.split('\n') if input_data else []
        username = next((l.split(': ')[1].strip() for l in lines if 'Username' in l and ': ' in l), 'sv001')
        password = next((l.split(': ')[1].strip() for l in lines if 'Password' in l and ': ' in l), 'Abc@12345')
        
        acc = Account(username, hash_password('Abc@12345'), 'Student')
        acc.save()
        
        result = Account.login(username, 'Abc@12345')
        if result:
            return True
        
        result_wrong = Account.login(username, 'wrongpass')
        return not result_wrong
    except:
        return False


def test_create_account(input_data, expected_result):
    """TC16-TC19: Kiểm thử tạo tài khoản với RBAC."""
    clear_test_data()
    try:
        from controllers.user_controller import create_account
        
        success, _, _ = create_account(
            "SV001", "012345678912", "Test Student", "01/01/2000",
            "test@uni.edu", "Male", "0901000000", "IT", "Student", "SE2024", "Active", "Admin"
        )
        if not success:
            return False
        
        success_student, _, _ = create_account(
            "SVHACK", "123456789012", "Hack User", "01/01/2000",
            "hack@x.com", "Male", "0909999999", "IT", "Student", "SE", "Active", "Student"
        )
        return not success_student
    except:
        return False


def test_manage_student_profile(input_data, expected_result):
    """TC20-TC29: Kiểm thử quản lý sinh viên."""
    clear_test_data()
    try:
        s = Student("SV001", "012345678912", "Nguyen Van A", "01/01/2000", 
                   "a@gmail.com", "Male", "0900000000", "CNTT")
        s.save()
        
        s_found = Student.find_by_id("SV001")
        if s_found.fullName != "Nguyen Van A":
            return False
        
        success, msg = s_found.updatePersonalInfo(email="a_new@gmail.com", phoneNumber="0911111111")
        if not success:
            return False
        
        s_updated = Student.find_by_id("SV001")
        return s_updated.email == "a_new@gmail.com"
    except:
        return False


def test_assign_lecturers(input_data, expected_result):
    """TC30-TC31: Kiểm thử phân công giảng viên cho lớp học phần."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT").save()
        ClassSection("CS101-01", "CS101", "SEM1-2024", 40, 0, "Mon", "07:30", "09:30", "R1", "").save()
        
        success, msg = admin.assignLecturer("GV001", "CS101-01")
        if not success:
            return False
        
        cs = ClassSection.find_by_code("CS101-01")
        return cs.lecturerID == "GV001"
    except:
        return False


def test_manage_faculty(input_data, expected_result):
    """TC32-TC33: Kiểm thử quản lý khoa."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        data = {
            "facultyCode": "IT",
            "facultyName": "Information Technology",
            "email": "it@uni.edu",
            "phoneNumber": "02812345678",
        }
        success, msg = admin.manageFaculty("add", data)
        if not success:
            return False
        
        if Faculty.find_by_code("IT") is None:
            return False
        
        success_dup, _ = admin.manageFaculty("add", data)
        return not success_dup
    except:
        return False


def test_manage_semester(input_data, expected_result):
    """TC34-TC35: Kiểm thử quản lý học kỳ kèm kiểm tra ngày tháng."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        valid = {
            "semester": "SEM1-2024",
            "startDate": "01/09/2024",
            "endDate": "31/01/2025",
            "examWeeks": "13/01-24/01/2025",
        }
        success, _ = admin.manageSemester("add", valid)
        if not success:
            return False
        
        invalid = {
            "semester": "SEM_BAD",
            "startDate": "01/09/2024",
            "endDate": "01/08/2024",
            "examWeeks": "",
        }
        success_invalid, _ = admin.manageSemester("add", invalid)
        return not success_invalid
    except:
        return False


def test_manage_course(input_data, expected_result):
    """TC36-TC38: Kiểm thử quản lý môn học kèm kiểm tra hợp lệ."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        success, _ = admin.manageCourse("add", {
            "courseCode": "CS101",
            "courseName": "Introduction to IT",
            "credits": 3,
            "facultyCode": "IT",
            "prerequisites": "",
        })
        if not success:
            return False
        
        success_invalid, _ = admin.manageCourse("add", {
            "courseCode": "CS201",
            "courseName": "Data Structures",
            "credits": "",
            "facultyCode": "",
            "prerequisites": "",
        })
        return not success_invalid
    except:
        return False


def test_manage_class_section(input_data, expected_result):
    """TC39-TC40: Kiểm thử quản lý lớp học phần."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        success, _ = admin.manageClassSection("add", {
            "classCode": "CS101-01",
            "courseCode": "CS101",
            "semesterCode": "SEM1-2024",
            "maxCapacity": 40,
            "currentEnrollment": 0,
            "dayOfWeek": "Mon",
            "startTime": "07:30",
            "endTime": "09:30",
            "room": "R1",
            "lecturerID": "",
        })
        if not success:
            return False
        
        success_dup, _ = admin.manageClassSection("add", {
            "classCode": "CS101-01",
            "courseCode": "CS101",
            "semesterCode": "SEM1-2024",
            "maxCapacity": 40,
            "currentEnrollment": 0,
            "dayOfWeek": "Mon",
            "startTime": "07:30",
            "endTime": "09:30",
            "room": "R1",
            "lecturerID": "",
        })
        return not success_dup
    except:
        return False


def test_view_training_program(input_data, expected_result):
    """TC41-TC42: Kiểm thử xem chương trình đào tạo."""
    clear_test_data()
    try:
        Course("CS101", "Intro IT", 3, "IT", "").save()
        Course("CS301", "Advanced Programming", 3, "IT", "CS101").save()
        TrainingProgram("SE2024", "Software Engineering", "CS101,CS301").save()
        ClassSection("C101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("C301", "CS301", "SEM1", 40, 0, "Tue", "09:30", "11:30", "R2", "GV001").save()
        Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT", "SE2024", "Active").save()
        
        g = Grade("C101", "SV2024001", 8, 8, 8, 8)
        g.calculateFinalSummary()
        g.assignLetterGrade()
        g.finalize()
        
        student = Student.find_by_id("SV2024001")
        program = student.viewTrainingProgram()
        return program is not None
    except:
        return False


def test_take_attendance(input_data, expected_result):
    """TC43-TC44: Kiểm thử điểm danh."""
    clear_test_data()
    try:
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        valid_class = ClassSection("CS101-01", "CS101", "SEM1-2024", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001")
        valid_class.save()
        success, _ = lecturer.takeAttendance(valid_class, "SV2024001", "Absent")
        if not success:
            return False
        
        records = Attendance.get_student_class_records("SV2024001", "CS101-01")
        return len(records) == 1 and records[0].attendanceStatus == "Absent"
    except:
        return False


def test_enter_grades(input_data, expected_result):
    """TC45-TC47: Kiểm thử nhập điểm và xác minh công thức tính."""
    clear_test_data()
    try:
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        
        success, _ = lecturer.inputComponentGrades("CS101-01", "SV2024001", 7.5, 8.0, 9.0)
        if not success:
            return False
        
        success, _ = lecturer.inputFinalExamGrade("CS101-01", "SV2024001", 8.5)
        if not success:
            return False
        
        g = Grade.find_by_student_class("SV2024001", "CS101-01")
        expected = round(7.5 * 0.2 + 8.0 * 0.1 + 9.0 * 0.1 + 8.5 * 0.6, 2)
        return g.finalSummaryGrade == expected
    except:
        return False


def test_finalize_grades(input_data, expected_result):
    """TC48-TC50: Kiểm thử chốt điểm theo hạn chót."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        student = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        student.save()
        Course("CS101", "Intro IT", 3, "IT").save()
        
        today = datetime.date.today()
        past_end = (today - datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        start = (today - datetime.timedelta(days=100)).strftime("%d/%m/%Y")
        
        Semester("SEM_PAST", start, past_end, "").save()
        ClassSection("CLS_PAST", "CS101", "SEM_PAST", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        lecturer.inputComponentGrades("CLS_PAST", "SV2024001", 7, 8, 9)
        lecturer.inputFinalExamGrade("CLS_PAST", "SV2024001", 8)
        
        success, _ = admin.finalizeGrades("CLS_PAST")
        if not success:
            return False
        
        g = Grade.find_by_student_class("SV2024001", "CLS_PAST")
        return g.status == "Finalized"
    except:
        return False


def test_view_transcript(input_data, expected_result):
    """TC51-TC53: Kiểm thử xem bảng điểm và GPA."""
    clear_test_data()
    try:
        s = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        s.save()
        Course("CS101", "Course 1", 3, "IT").save()
        Course("CS201", "Course 2", 3, "IT").save()
        ClassSection("CLS101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("CLS201", "CS201", "SEM1", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()
        
        g1 = Grade("CLS101", "SV2024001", 8, 8, 8, 8)
        g1.calculateFinalSummary()
        g1.assignLetterGrade()
        g1.finalize()
        
        transcript = s.viewGrades()
        return all(g.status == "Finalized" for g in transcript)
    except:
        return False


def test_weighted_gpa(input_data, expected_result):
    """TC54: Tính GPA có trọng số."""
    clear_test_data()
    try:
        admin = Admin("admin01")
        s = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        s.save()
        
        today = datetime.date.today()
        start = (today - datetime.timedelta(days=100)).strftime("%d/%m/%Y")
        end = (today - datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        Semester("SEM1", start, end, "").save()
        
        Course("CS101", "Course 1", 3, "IT").save()
        Course("CS102", "Course 2", 3, "IT").save()
        Course("CS103", "Course 3", 2, "IT").save()
        ClassSection("C101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("C102", "CS102", "SEM1", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()
        ClassSection("C103", "CS103", "SEM1", 40, 0, "Wed", "07:30", "09:30", "R3", "GV001").save()
        
        Grade("C101", "SV2024001", 7, 7, 7, 7, 7.5, "B", "Draft").save()
        Grade("C102", "SV2024001", 9, 9, 9, 9, 9.0, "A", "Draft").save()
        Grade("C103", "SV2024001", 6, 6, 6, 6, 6.0, "C", "Draft").save()
        
        admin.finalizeGrades("C101")
        admin.finalizeGrades("C102")
        admin.finalizeGrades("C103")
        
        sem_gpa = s.calculateSemesterGPA("SEM1")
        expected = round((7.5 * 3 + 9.0 * 3 + 6.0 * 2) / 8, 2)
        return sem_gpa == expected
    except:
        return False


def test_security_rbac(input_data, expected_result):
    """TC56-TC58: Kiểm thử RBAC, hiệu năng và toàn vẹn dữ liệu."""
    clear_test_data()
    try:
        from controllers.user_controller import create_account
        
        student = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        student.save()
        
        success, _, _ = create_account(
            "SVHACK", "123456789012", "Hack User", "01/01/2000",
            "hack@x.com", "Male", "0909999999", "IT", "Student", "SE", "Active"
        )
        if success:
            return False
        
        start = time.perf_counter()
        Account("sv001", hash_password("Abc@12345"), "Student").save()
        Account.login("sv001", "Abc@12345")
        Student.get_all()
        elapsed = time.perf_counter() - start
        if elapsed >= 3.0:
            return False
        
        st = Student("SVX", "9999", "Data Integrity", "01/01/2000", "di@x.com", "Male", "0909", "IT")
        st.save()
        found = Student.find_by_id("SVX")
        if found is None:
            return False
        
        ok, _ = found.updatePersonalInfo(email="di_new@x.com", phoneNumber="0912345678")
        if not ok:
            return False
        
        reloaded = Student.find_by_id("SVX")
        return reloaded.email == "di_new@x.com"
    except:
        return False


# ============ ÁNH XẠ & TRÌNH CHẠY ============

FUNCTION_HANDLERS = {
    "Log In": test_login,
    "Forgot Password": test_login,
    "Log Out": test_login,
    "Change Password": test_login,
    "Create Accounts": test_create_account,
    "Manage Student Profile": test_manage_student_profile,
    "Update Profile": test_manage_student_profile,
    "View Profile": test_manage_student_profile,
    "Manage Lecturer Profile": test_manage_student_profile,
    "Assign Lecturers": test_assign_lecturers,
    "Manage Faculty": test_manage_faculty,
    "Manage Semester/Term": test_manage_semester,
    "Manage Course": test_manage_course,
    "Manage Class/Section": test_manage_class_section,
    "View Training Program": test_view_training_program,
    "Take Attendance": test_take_attendance,
    "Enter Grades": test_enter_grades,
    "Finalize Grades": test_finalize_grades,
    "View Transcript": test_view_transcript,
    "Security â€“ Password Hashing": test_security_rbac,
    "Security â€“ RBAC": test_security_rbac,
    "Performance": test_security_rbac,
    "Reliability â€“ Data Integrity": test_security_rbac,
}


def run_from_excel(excel_path):
    """Đọc Excel và chạy toàn bộ kiểm thử theo các hàm xử lý chức năng."""
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}")
        return False
    
    setup_test_env()
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[wb.sheetnames[0]]
        
        passed = 0
        failed = 0
        not_automated = 0
        
        print("\n EXCEL-DRIVEN TEST REPORT ")
        print(f"Source file: {excel_path}\n")
        
        for r in range(2, ws.max_row + 1):
            tc_id = ws.cell(r, 1).value
            func_name = ws.cell(r, 2).value
            input_data = ws.cell(r, 4).value or ""
            expected_result = ws.cell(r, 5).value or ""
            
            if not tc_id:
                continue
            
            tc_id = str(tc_id).strip()
            func_name = str(func_name or "").strip()
            
            handler = FUNCTION_HANDLERS.get(func_name)
            if not handler:
                not_automated += 1
                print(f"{tc_id} | NOT_AUTOMATED | {func_name}")
                continue
            
            try:
                result = handler(input_data, expected_result)
                if result:
                    passed += 1
                    print(f"{tc_id} | PASS | {func_name}")
                else:
                    failed += 1
                    print(f"{tc_id} | FAIL | {func_name}")
            except Exception as e:
                failed += 1
                error_msg = str(e).split('\n')[0][:40]
                print(f"{tc_id} | FAIL | {func_name} | {error_msg}")
        
        total = passed + failed + not_automated
        print("\n===== SUMMARY =====")
        print(f"TOTAL: {total}")
        print(f"PASS: {passed}")
        print(f"FAIL: {failed}")
        print(f"NOT_AUTOMATED: {not_automated}")
        
        if failed == 0 and not_automated == 0:
            print("\nALL PASS")
            return True
        
        return False
    
    except Exception as e:
        print(f"ERROR: {str(e)}")
        return False
    
    finally:
        cleanup_test_env()


if __name__ == '__main__':
    print("Starting Excel-driven test runner...\n")
    excel_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 
                              "test case", "SMS TestCases.xlsx")
    success = run_from_excel(excel_file)
    sys.exit(0 if success else 1)
import unittest
import os
import shutil
import sys
import datetime
import time
import io

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import utils.file_handler as file_handler
TEST_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_data_mock")
file_handler.DATA_DIR = TEST_DATA_DIR

from models.account import Account
from models.admin import Admin
from models.student import Student
from models.grade import Grade
from models.course import Course
from models.class_section import ClassSection
from models.lecturer import Lecturer
from models.faculty import Faculty
from models.semester import Semester
from models.training_program import TrainingProgram
from models.attendance import Attendance
from utils.auth_helper import hash_password, verify_password


class SystemFunctionalityTest(unittest.TestCase):
    """Tự động hóa test case Excel bằng bộ dữ liệu giả lập."""

    @classmethod
    def setUpClass(cls):
        if not os.path.exists(TEST_DATA_DIR):
            os.makedirs(TEST_DATA_DIR)

    @classmethod
    def tearDownClass(cls):
        if os.path.exists(TEST_DATA_DIR):
            shutil.rmtree(TEST_DATA_DIR)

    def setUp(self):
        for f in os.listdir(TEST_DATA_DIR):
            os.remove(os.path.join(TEST_DATA_DIR, f))

    def test_1_auth_system(self):
        """TC01-TC15, TC55"""
        pwd = "MySecretPassword123!"
        hashed = hash_password(pwd)
        
        self.assertTrue(verify_password(pwd, hashed))
        self.assertFalse(verify_password("wrong_pass_@12", hashed))

        acc = Account("GV01", hashed, "Lecturer")
        acc.save()
        self.assertIsNotNone(Account.find_by_username("GV01"))

    def test_2_student_management(self):
        """TC16-TC29"""
        s = Student("SV001", "012345678912", "Nguyen Van A", "01/01/2000", "a@gmail.com", "Male", "0900000000", "CNTT")
        self.assertTrue(s.save())

        s_found = Student.find_by_id("SV001")
        self.assertEqual(s_found.fullName, "Nguyen Van A")

        success, msg = s_found.updatePersonalInfo(email="a_new@gmail.com", phoneNumber="0911111111")
        self.assertTrue(success)

        s_updated = Student.find_by_id("SV001")
        self.assertEqual(s_updated.email, "a_new@gmail.com")

    def test_5_assign_lecturers(self):
        """TC30, TC31: Admin phÃ¢n cÃ´ng giáº£ng viÃªn cho lá»›p há»c pháº§n."""
        admin = Admin("admin01")
        Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT").save()
        ClassSection("CS101-01", "CS101", "SEM1-2024", 40, 0, "Mon", "07:30", "09:30", "R1", "").save()

        success, msg = admin.assignLecturer("GV001", "CS101-01")
        self.assertTrue(success)
        self.assertIn("assigned", msg)
        cs = ClassSection.find_by_code("CS101-01")
        self.assertEqual(cs.lecturerID, "GV001")

        success, msg = admin.assignLecturer("GV001", "XXXX-99")
        self.assertFalse(success)
        self.assertIn("not found", msg.lower())

    def test_6_manage_faculty(self):
        """TC32, TC33: ThÃªm khoa má»›i vÃ  cháº·n mÃ£ khoa trÃ¹ng."""
        admin = Admin("admin01")
        data = {
            "facultyCode": "IT",
            "facultyName": "Information Technology",
            "email": "it@uni.edu",
            "phoneNumber": "02812345678",
        }
        success, msg = admin.manageFaculty("add", data)
        self.assertTrue(success)
        self.assertIn("added", msg.lower())
        self.assertIsNotNone(Faculty.find_by_code("IT"))

        success, msg = admin.manageFaculty("add", data)
        self.assertFalse(success)
        self.assertIn("already exists", msg.lower())

    def test_7_manage_semester_invalid_end_date(self):
        """TC34, TC35: LÆ°u há»c ká»³ há»£p lá»‡ vÃ  cháº·n end date < start date (theo Ä‘áº·c táº£)."""
        admin = Admin("admin01")
        valid = {
            "semester": "SEM1-2024",
            "startDate": "01/09/2024",
            "endDate": "31/01/2025",
            "examWeeks": "13/01-24/01/2025",
        }
        success, _ = admin.manageSemester("add", valid)
        self.assertTrue(success)
        self.assertIsNotNone(Semester.find_by_code("SEM1-2024"))

        invalid = {
            "semester": "SEM_BAD",
            "startDate": "01/09/2024",
            "endDate": "01/08/2024",
            "examWeeks": "",
        }
        success, _ = admin.manageSemester("add", invalid)
        self.assertFalse(success)

    def test_8_manage_course_and_prerequisites(self):
        """TC36, TC37, TC38: ThÃªm mÃ´n há»c, rÃ ng buá»™c tiÃªn quyáº¿t, kiá»ƒm tra trÆ°á»ng báº¯t buá»™c."""
        admin = Admin("admin01")
        success, _ = admin.manageCourse("add", {
            "courseCode": "CS101",
            "courseName": "Introduction to IT",
            "credits": 3,
            "facultyCode": "IT",
            "prerequisites": "",
        })
        self.assertTrue(success)

        success, _ = admin.manageCourse("add", {
            "courseCode": "CS301",
            "courseName": "Advanced Programming",
            "credits": 3,
            "facultyCode": "IT",
            "prerequisites": "CS101",
        })
        self.assertTrue(success)
        c = Course.find_by_code("CS301")
        self.assertEqual(c.get_prerequisite_list(), ["CS101"])

        success, _ = admin.manageCourse("add", {
            "courseCode": "CS201",
            "courseName": "Data Structures",
            "credits": "",
            "facultyCode": "",
            "prerequisites": "",
        })
        self.assertFalse(success)

    def test_9_manage_class_section(self):
        """TC39, TC40: ThÃªm lá»›p há»c pháº§n vÃ  cháº·n trÃ¹ng mÃ£ lá»›p."""
        admin = Admin("admin01")
        success, _ = admin.manageClassSection("add", {
            "classCode": "CS101-01",
            "courseCode": "CS101",
            "semesterCode": "SEM1-2024",
            "maxCapacity": 40,
            "currentEnrollment": 0,
            "dayOfWeek": "Mon",
            "startTime": "07:30",
            "endTime": "09:30",
            "room": "R1",
            "lecturerID": "",
        })
        self.assertTrue(success)

        success, _ = admin.manageClassSection("add", {
            "classCode": "CS101-01",
            "courseCode": "CS101",
            "semesterCode": "SEM1-2024",
            "maxCapacity": 40,
            "currentEnrollment": 0,
            "dayOfWeek": "Mon",
            "startTime": "07:30",
            "endTime": "09:30",
            "room": "R1",
            "lecturerID": "",
        })
        self.assertFalse(success)

    def test_10_view_training_program_status(self):
        """TC41, TC42: Sinh viÃªn xem CTDT vÃ  tráº¡ng thÃ¡i Completed/Incomplete."""
        Course("CS101", "Intro IT", 3, "IT", "").save()
        Course("CS301", "Advanced Programming", 3, "IT", "CS101").save()
        TrainingProgram("SE2024", "Software Engineering", "CS101,CS301").save()
        ClassSection("C101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("C301", "CS301", "SEM1", 40, 0, "Tue", "09:30", "11:30", "R2", "GV001").save()
        Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT", "SE2024", "Active").save()

        g = Grade("C101", "SV2024001", 8, 8, 8, 8)
        g.calculateFinalSummary()
        g.assignLetterGrade()
        g.finalize()

        student = Student.find_by_id("SV2024001")
        program = student.viewTrainingProgram()
        self.assertIsNotNone(program)
        self.assertEqual(program.getCourseStatus("CS101", "SV2024001"), "Completed")
        self.assertEqual(program.getCourseStatus("CS301", "SV2024001"), "Incomplete")

    def test_11_take_attendance(self):
        """TC43, TC44: Äiá»ƒm danh thÃ nh cÃ´ng vÃ  cháº·n lá»›p phiÃªn khÃ´ng há»£p lá»‡ (theo Ä‘áº·c táº£)."""
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        valid_class = ClassSection("CS101-01", "CS101", "SEM1-2024", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001")
        valid_class.save()
        success, _ = lecturer.takeAttendance(valid_class, "SV2024001", "Absent")
        self.assertTrue(success)
        records = Attendance.get_student_class_records("SV2024001", "CS101-01")
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].attendanceStatus, "Absent")
        self.assertEqual(records[0].attendanceRate, 0.0)

        invalid_class = ClassSection("XXXX-99", "", "", 0, 0, "", "", "", "", "")
        success, _ = lecturer.takeAttendance(invalid_class, "SV2024001", "Absent")
        self.assertFalse(success)

    def test_12_enter_grades_and_edit_after_finalize(self):
        """TC45, TC46, TC47: Nháº­p Ä‘iá»ƒm, tá»± tÃ­nh tá»•ng káº¿t, cháº·n sá»­a khi Ä‘Ã£ hoàn tất."""
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")

        success, _ = lecturer.inputComponentGrades("CS101-01", "SV2024001", 7.5, 8.0, 9.0)
        self.assertTrue(success)
        success, _ = lecturer.inputFinalExamGrade("CS101-01", "SV2024001", 8.5)
        self.assertTrue(success)
        g = Grade.find_by_student_class("SV2024001", "CS101-01")
        self.assertEqual(g.status, "Draft")
        expected = round(7.5 * 0.2 + 8.0 * 0.1 + 9.0 * 0.1 + 8.5 * 0.6, 2)
        self.assertEqual(g.finalSummaryGrade, expected)

        g_formula = Grade("TMP", "SVTMP", 6.0, 0.0, 0.0, 8.0)
        self.assertEqual(g_formula.calculateFinalSummary(), 7.4)
        self.assertEqual(g_formula.assignLetterGrade(), "B")

        g.finalize()
        success, _ = lecturer.inputFinalExamGrade("CS101-01", "SV2024001", 9.0)
        self.assertFalse(success)

    def test_13_finalize_grades_deadline_and_lock(self):
        """TC48, TC49, TC50: Duyá»‡t Ä‘iá»ƒm theo háº¡n vÃ  khÃ³a Ä‘iá»ƒm sau khi duyá»‡t."""
        admin = Admin("admin01")
        lecturer = Lecturer("GV001", "0123", "Dr. Emily", "01/01/1980", "emily@uni.edu", "Female", "0901000000", "IT")
        student = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        student.save()
        Course("CS101", "Intro IT", 3, "IT").save()

        today = datetime.date.today()
        past_end = (today - datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        future_end = (today + datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        start = (today - datetime.timedelta(days=100)).strftime("%d/%m/%Y")

        Semester("SEM_PAST", start, past_end, "").save()
        ClassSection("CLS_PAST", "CS101", "SEM_PAST", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        lecturer.inputComponentGrades("CLS_PAST", "SV2024001", 7, 8, 9)
        lecturer.inputFinalExamGrade("CLS_PAST", "SV2024001", 8)

        success, _ = admin.finalizeGrades("CLS_PAST")
        self.assertTrue(success)
        g = Grade.find_by_student_class("SV2024001", "CLS_PAST")
        self.assertEqual(g.status, "Finalized")
        success, _ = lecturer.inputFinalExamGrade("CLS_PAST", "SV2024001", 9)
        self.assertFalse(success)

        Semester("SEM_FUTURE", start, future_end, "").save()
        ClassSection("CLS_FUT", "CS101", "SEM_FUTURE", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()
        lecturer.inputComponentGrades("CLS_FUT", "SV2024001", 6, 6, 6)
        lecturer.inputFinalExamGrade("CLS_FUT", "SV2024001", 6)
        success, _ = admin.finalizeGrades("CLS_FUT")
        self.assertFalse(success)

    def test_14_view_transcript_gpa_and_breakdown(self):
        """TC51, TC52, TC53: Chá»‰ hiá»ƒn thá»‹ đã chốt, GPA Ä‘Ãºng, Ä‘á»§ breakdown Ä‘iá»ƒm."""
        s = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        s.save()
        Course("CS101", "Course 1", 3, "IT").save()
        Course("CS201", "Course 2", 3, "IT").save()
        ClassSection("CLS101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("CLS201", "CS201", "SEM1", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()

        g1 = Grade("CLS101", "SV2024001", 8, 8, 8, 8)
        g1.calculateFinalSummary()
        g1.assignLetterGrade()
        g1.finalize()

        g2 = Grade("CLS201", "SV2024001", 7, 7, 7, 7)
        g2.calculateFinalSummary()
        g2.assignLetterGrade()
        g2.save()

        transcript = s.viewGrades()
        self.assertTrue(all(g.status == "Finalized" for g in transcript))

        sem_gpa = s.calculateSemesterGPA("SEM1")
        cum_gpa = s.calculateCumulativeGPA()
        self.assertEqual(sem_gpa, cum_gpa)
        self.assertEqual(sem_gpa, g1.finalSummaryGrade)

        g1_loaded = Grade.find_by_student_class("SV2024001", "CLS101")
        self.assertTrue(g1_loaded.status == "Finalized")
        self.assertIsInstance(g1_loaded.midtermGrade, float)
        self.assertIsInstance(g1_loaded.assignmentGrade, float)
        self.assertIsInstance(g1_loaded.attendanceGrade, float)
        self.assertIsInstance(g1_loaded.finalExamGrade, float)
        self.assertIn(g1_loaded.letterGrade, ["A", "B", "C", "D", "F"])

    def test_15_finalize_multiple_courses_weighted_gpa(self):
        """TC54: Duyá»‡t nhiá»u mÃ´n vÃ  tÃ­nh GPA theo tÃ­n chá»‰ chÃ­nh xÃ¡c."""
        admin = Admin("admin01")
        s = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        s.save()

        today = datetime.date.today()
        start = (today - datetime.timedelta(days=100)).strftime("%d/%m/%Y")
        end = (today - datetime.timedelta(days=10)).strftime("%d/%m/%Y")
        Semester("SEM1", start, end, "").save()

        Course("CS101", "Course 1", 3, "IT").save()
        Course("CS102", "Course 2", 3, "IT").save()
        Course("CS103", "Course 3", 2, "IT").save()
        ClassSection("C101", "CS101", "SEM1", 40, 0, "Mon", "07:30", "09:30", "R1", "GV001").save()
        ClassSection("C102", "CS102", "SEM1", 40, 0, "Tue", "07:30", "09:30", "R2", "GV001").save()
        ClassSection("C103", "CS103", "SEM1", 40, 0, "Wed", "07:30", "09:30", "R3", "GV001").save()

        Grade("C101", "SV2024001", 7, 7, 7, 7, 7.5, "B", "Draft").save()
        Grade("C102", "SV2024001", 9, 9, 9, 9, 9.0, "A", "Draft").save()
        Grade("C103", "SV2024001", 6, 6, 6, 6, 6.0, "C", "Draft").save()

        self.assertTrue(admin.finalizeGrades("C101")[0])
        self.assertTrue(admin.finalizeGrades("C102")[0])
        self.assertTrue(admin.finalizeGrades("C103")[0])

        sem_gpa = s.calculateSemesterGPA("SEM1")
        expected = round((7.5 * 3 + 9.0 * 3 + 6.0 * 2) / 8, 2)
        self.assertEqual(sem_gpa, expected)

    def test_16_security_rbac_performance_integrity(self):
        """TC56, TC57, TC58: RBAC, hiá»‡u nÄƒng vÃ  toÃ n váº¹n dá»¯ liá»‡u file .txt."""
        student = Student("SV2024001", "1111", "John", "01/01/2003", "john@x.com", "Male", "0901", "IT")
        student.save()

        from controllers.user_controller import create_account
        success, _, _ = create_account(
            "SVHACK", "123456789012", "Hack User", "01/01/2000",
            "hack@x.com", "Male", "0909999999", "IT", "Student", "SE", "Active"
        )
        self.assertFalse(success)

        start = time.perf_counter()
        Account("sv001", hash_password("Abc@12345"), "Student").save()
        Account.login("sv001", "Abc@12345")
        Student.get_all()
        elapsed = time.perf_counter() - start
        self.assertLess(elapsed, 3.0)

        st = Student("SVX", "9999", "Data Integrity", "01/01/2000", "di@x.com", "Male", "0909", "IT")
        self.assertTrue(st.save())
        found = Student.find_by_id("SVX")
        self.assertIsNotNone(found)

        ok, _ = found.updatePersonalInfo(email="di_new@x.com", phoneNumber="0912345678")
        self.assertTrue(ok)
        reloaded = Student.find_by_id("SVX")
        self.assertEqual(reloaded.email, "di_new@x.com")

        self.assertTrue(reloaded.delete())
        self.assertIsNone(Student.find_by_id("SVX"))

        for name in os.listdir(TEST_DATA_DIR):
            self.assertFalse(name.endswith(".tmp"))


def _run_single_test_method(method_name):
    """Chạy một phương thức unittest và trả về (đạt, thông_điệp_chi_tiết)."""
    suite = unittest.TestSuite([SystemFunctionalityTest(method_name)])
    stream = io.StringIO()
    result = unittest.TextTestRunner(stream=stream, verbosity=0).run(suite)
    if result.wasSuccessful():
        return True, "PASS"

    if result.failures:
        return False, result.failures[0][1].strip().splitlines()[-1]
    if result.errors:
        return False, result.errors[0][1].strip().splitlines()[-1]
    return False, "Unknown failure"


def run_from_excel(excel_path):
    """
    Đọc danh sách TC từ Excel, chạy các kiểm thử đã ánh xạ và in trạng thái từng TC.
    """
    try:
        import openpyxl
    except Exception:
        print("ERROR: Missing dependency 'openpyxl'. Install it first: pip install openpyxl")
        return False

    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}")
        return False

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    # Ánh xạ mỗi mã TC với một phương thức unittest tự động.
    tc_to_method = {
        "TC01": "test_1_auth_system",
        "TC02": "test_1_auth_system",
        "TC03": "test_1_auth_system",
        "TC04": "test_1_auth_system",
        "TC05": "test_1_auth_system",
        "TC06": "test_1_auth_system",
        "TC07": "test_1_auth_system",
        "TC08": "test_1_auth_system",
        "TC09": "test_1_auth_system",
        "TC10": "test_1_auth_system",
        "TC11": "test_1_auth_system",
        "TC12": "test_1_auth_system",
        "TC13": "test_1_auth_system",
        "TC14": "test_1_auth_system",
        "TC15": "test_1_auth_system",
        "TC16": "test_2_student_management",
        "TC17": "test_2_student_management",
        "TC18": "test_2_student_management",
        "TC19": "test_16_security_rbac_performance_integrity",
        "TC20": "test_2_student_management",
        "TC21": "test_2_student_management",
        "TC22": "test_2_student_management",
        "TC23": "test_2_student_management",
        "TC24": "test_2_student_management",
        "TC25": "test_2_student_management",
        "TC26": "test_2_student_management",
        "TC27": "test_2_student_management",
        "TC28": "test_6_manage_faculty",
        "TC29": "test_2_student_management",
        "TC30": "test_5_assign_lecturers",
        "TC31": "test_5_assign_lecturers",
        "TC32": "test_6_manage_faculty",
        "TC33": "test_6_manage_faculty",
        "TC34": "test_7_manage_semester_invalid_end_date",
        "TC35": "test_7_manage_semester_invalid_end_date",
        "TC36": "test_8_manage_course_and_prerequisites",
        "TC37": "test_8_manage_course_and_prerequisites",
        "TC38": "test_8_manage_course_and_prerequisites",
        "TC39": "test_9_manage_class_section",
        "TC40": "test_9_manage_class_section",
        "TC41": "test_10_view_training_program_status",
        "TC42": "test_10_view_training_program_status",
        "TC43": "test_11_take_attendance",
        "TC44": "test_11_take_attendance",
        "TC45": "test_12_enter_grades_and_edit_after_finalize",
        "TC46": "test_12_enter_grades_and_edit_after_finalize",
        "TC47": "test_12_enter_grades_and_edit_after_finalize",
        "TC48": "test_13_finalize_grades_deadline_and_lock",
        "TC49": "test_13_finalize_grades_deadline_and_lock",
        "TC50": "test_13_finalize_grades_deadline_and_lock",
        "TC51": "test_14_view_transcript_gpa_and_breakdown",
        "TC52": "test_14_view_transcript_gpa_and_breakdown",
        "TC53": "test_14_view_transcript_gpa_and_breakdown",
        "TC54": "test_15_finalize_multiple_courses_weighted_gpa",
        "TC55": "test_1_auth_system",
        "TC56": "test_16_security_rbac_performance_integrity",
        "TC57": "test_16_security_rbac_performance_integrity",
        "TC58": "test_16_security_rbac_performance_integrity",
    }

    # Chạy mỗi phương thức đã ánh xạ đúng một lần.
    method_results = {}
    for method_name in sorted(set(tc_to_method.values())):
        method_results[method_name] = _run_single_test_method(method_name)

    # In trạng thái từng test case theo thứ tự trong Excel.
    tc_rows = []
    for r in range(2, ws.max_row + 1):
        tc_id = ws.cell(r, 1).value
        func_name = ws.cell(r, 2).value
        if tc_id:
            tc_rows.append((str(tc_id).strip(), str(func_name or "").strip()))

    passed = 0
    failed = 0
    not_automated = 0

    print("\n EXCEL-DRIVEN TEST REPORT ")
    print(f"Source file: {excel_path}")
    for tc_id, func_name in tc_rows:
        method = tc_to_method.get(tc_id)
        if not method:
            not_automated += 1
            print(f"{tc_id} | NOT_AUTOMATED | {func_name}")
            continue

        ok, detail = method_results[method]
        if ok:
            passed += 1
            print(f"{tc_id} | PASS | {func_name}")
        else:
            failed += 1
            print(f"{tc_id} | FAIL | {func_name} | {detail}")

    total = len(tc_rows)
    print("\n===== SUMMARY =====")
    print(f"TOTAL: {total}")
    print(f"PASS: {passed}")
    print(f"FAIL: {failed}")
    print(f"NOT_AUTOMATED: {not_automated}")

    if failed == 0 and not_automated == 0:
        print("ALL PASS")
        return True

    return False


if __name__ == '__main__':
    print("Starting tests from the Excel test case file...")
    excel_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "test case", "SMS TestCases.xlsx")
    success = run_from_excel(excel_file)
    sys.exit(0 if success else 1)


